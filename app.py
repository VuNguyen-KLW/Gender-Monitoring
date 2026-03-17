from flask import Flask, request, jsonify, render_template, send_file
from io import BytesIO
from openpyxl import Workbook, load_workbook
import json
import os

app = Flask(__name__)
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")


def read_data():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r") as f:
        return json.load(f)


def write_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)


@app.route("/")
def form_page():
    return render_template("form.html")


@app.route("/dashboard")
def dashboard_page():
    return render_template("dashboard.html")


@app.route("/api/users", methods=["GET"])
def get_users():
    users = read_data()
    gender = request.args.get("gender")
    age_min = request.args.get("age_min", type=int)
    age_max = request.args.get("age_max", type=int)

    if gender:
        users = [u for u in users if u["gender"] == gender]
    if age_min is not None:
        users = [u for u in users if u["age"] >= age_min]
    if age_max is not None:
        users = [u for u in users if u["age"] <= age_max]

    return jsonify(users)


@app.route("/api/users", methods=["POST"])
def add_user():
    data = request.get_json()
    if not data or not all(k in data for k in ("name", "age", "gender")):
        return jsonify({"error": "name, age, and gender are required"}), 400

    try:
        age = int(data["age"])
    except (ValueError, TypeError):
        return jsonify({"error": "age must be a number"}), 400

    user = {"name": data["name"].strip(), "age": age, "gender": data["gender"]}
    users = read_data()
    users.append(user)
    write_data(users)
    return jsonify({"message": "User added successfully"}), 201


@app.route("/api/users/template", methods=["GET"])
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Name", "Age", "Gender"])
    ws.append(["John Doe", 30, "Male"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="user_template.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/users/bulk", methods=["POST"])
def bulk_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are accepted"}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        return jsonify({"error": "Invalid Excel file"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "Empty file"}), 400

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map = {}
    for i, h in enumerate(header):
        if h == "name":
            col_map["name"] = i
        elif h == "age":
            col_map["age"] = i
        elif h == "gender":
            col_map["gender"] = i

    missing = {"name", "age", "gender"} - set(col_map.keys())
    if missing:
        return jsonify({"error": f"Missing columns: {', '.join(missing)}"}), 400

    gender_map = {
        "male": "Male", "female": "Female", "other": "Other",
        "nam": "Nam", "nữ": "Nữ",
    }
    new_users = []
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_errors = []
        name = row[col_map["name"]]
        age = row[col_map["age"]]
        gender = row[col_map["gender"]]

        if not name or str(name).strip() == "":
            row_errors.append("Name is required")
        else:
            name = str(name).strip()

        try:
            age = int(age)
            if age < 0 or age > 150:
                row_errors.append("Age must be between 0 and 150")
        except (ValueError, TypeError):
            row_errors.append("Age must be a valid number")
            age = None

        gender_key = str(gender).strip().lower() if gender else ""
        if gender_key not in gender_map:
            row_errors.append("Gender must be Male/Female/Other or Nam/Nữ")
        else:
            gender = gender_map[gender_key]

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})
        else:
            new_users.append({"name": name, "age": age, "gender": gender})

    if new_users:
        users = read_data()
        users.extend(new_users)
        write_data(users)

    return jsonify({
        "success_count": len(new_users),
        "error_count": len(errors),
        "errors": errors
    })


if __name__ == "__main__":
    app.run(debug=True)
